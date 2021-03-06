# @Author: 夭夭
# @Time: 五月 28, 2020


import warnings
import time
from keywordLibrary.keywordFunction_pre import KeyWord

from tool.element import ELEMENT
import ddt
import unittest
from tool.readtestcase_openpyxl import CreateExcel
# 导入字体、边框、颜色以及对齐方式相关库
from openpyxl.styles import PatternFill

# 为了回写用的填充样式
red_fill = PatternFill("solid", fgColor="FF0000")
green_fill = PatternFill("solid", fgColor="00FF00")
filezpath = "mtxshop.xlsx"
at = CreateExcel(filezpath)
# 获取工作簿wb
wb = at.get_book()
# 获取要执行的测试用例
case_list = at.read_case_y_n()

at.read_element()
# 获取测试元素库的数据
# ws = wb['测试元素库']
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row,min_col=1,max_col=3):
#     col_value = [cell.value for cell in row]
#     # 登录链接 = xpath,//div[@class='member-login']/a[text()='登录']
#     ELEMENT[col_value[0]] = (col_value[1], col_value[2])
#     print('ELEMENT的值是', ELEMENT)



@ddt.ddt
class TestRunScript(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.kwFunc = KeyWord()

    @ddt.data(*case_list)
    def test_case(self, case):

        print(case)
        report_result = []
        # 获取到登录sheet页-ws1
        ws1 = wb[case[0]]
        # row代表的是行数，从第2行开始，iter_rows迭代结束，证明行数结束
        row = 2
        for i in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=4):
            data = [cell.value for cell in i]
            print('data的值是', data)
            if data[0] == '打开':
                try:
                    self.kwFunc.base_open(url=data[2])
                    ws1.cell(row, 4, '测试成功').fill = green_fill
                    report_result.append('success')
                except:
                    ws1.cell(row, 4, '测试失败').fill = red_fill
                    report_result.append('fail')
            elif data[0] == '点击':
                try:
                    self.kwFunc.base_click(data[1])
                    ws1.cell(row, 4, '测试成功').fill = green_fill
                    report_result.append('success')
                except:
                    ws1.cell(row, 4, '测试失败').fill = red_fill
                    report_result.append('fail')
            elif data[0] == '输入':
                try:
                    self.kwFunc.base_input(data[1], data[2])
                    ws1.cell(row, 4, '测试成功').fill = green_fill
                    report_result.append('success')
                except:
                    ws1.cell(row, 4, '测试失败').fill = red_fill
                    report_result.append('fail')

            elif data[0] == '等待':
                try:
                    self.kwFunc.base_wait(int(data[2]))
                    ws1.cell(row, 4, '测试成功').fill = green_fill
                    report_result.append('success')
                except:
                    ws1.cell(row, 4, '测试失败').fill = red_fill
                    report_result.append('fail')
            elif data[0] == '断言':
                try:
                    # result指的就是page_source页面源码，断言就是看某个数据是否在页面源码里面
                    result = self.kwFunc.base_assert_contain_string()
                    self.assertIn(data[2], result)
                    ws1.cell(row, 4, '测试成功').fill = green_fill
                    report_result.append('success')
                except:
                    ws1.cell(row, 4, '测试失败').fill = red_fill
                    report_result.append('fail')
            elif data[0] == '切换window':
                try:
                    self.kwFunc.base_switch_window(data[2])
                    ws1.cell(row, 4, '测试成功').fill = green_fill
                    report_result.append('success')
                except:
                    ws1.cell(row, 4, '测试失败').fill = red_fill
                    report_result.append('fail')
            print('report_result是',report_result)
            row += 1

            # 调用保存数据的方法

        wb.save('../data/mtxshop.xlsx')
        print('保存成功')
        ws2 = wb['测试用例集合']
        # 先获取case在整个case_list里面的索引值，+2就可以知道他的行数
        # 因为读取excel就是按照用例的行数一个一个去读取的
        index = case_list.index(case)
        print(f'index的值是{index}')
        if 'fail' in report_result:
            ws2.cell(index+2, 4, '测试失败').fill = red_fill
        else:
            ws2.cell(index+2, 4, '测试成功').fill = green_fill
            print('我被执行了')
        time.sleep(3)
        # 调用保存数据的方法
        wb.save('../data/mtxshop.xlsx')

        print('保存成功')
    @classmethod
    def tearDownClass(cls):
        cls.kwFunc.base_quit_driver()


if __name__ == '__main__':
    unittest.main()