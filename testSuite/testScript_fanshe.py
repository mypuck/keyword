# @Author: 夭夭
# @Time: 五月 28, 2020


import warnings
import time
from keywordLibrary.keywordFunction_fanshe import KeyWord

from tool.element import ELEMENT
import ddt
import unittest
from tool.readtestcase_openpyxl import CreateExcel
# 导入字体、边框、颜色以及对齐方式相关库
from openpyxl.styles import PatternFill

# 为了回写用的填充样式
red_fill = PatternFill("solid", fgColor="FF0000")
green_fill = PatternFill("solid", fgColor="00FF00")
filezpath = "mtxshop.xlsx"
at = CreateExcel(filezpath)
# 获取工作簿wb
wb = at.get_book()
# 获取要执行的测试用例
case_list = at.read_case()

at.read_element()


@ddt.ddt
class TestRunScript(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.kwFunc = KeyWord()

    @ddt.data(*case_list)
    def test_case(self, case):
        dic = {"打开":'base_open', '点击':'base_click',
                '输入':'base_input', '等待':'base_wait','断言':'base_assert_contain_string',
                '切换window':'base_switch_window'}
        print(case)
        report_result = []
        # 获取到登录sheet页-ws1
        ws1 = wb[case[0]]
        # row代表的是行数，从第2行开始，iter_rows迭代结束，证明行数结束
        row = 2
        for i in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=4):
            data = [cell.value for cell in i]
            print('data的值是', data)
            try:
                # 方法名字 add                      ()
                # 对于函数的参数不一样的问题
                # 解决方法：都让他传一样的参数，
                getattr(self.kwFunc, dic[data[0]])(data[1],data[2])
                ws1.cell(row, 4, '测试成功').fill = green_fill
                report_result.append('success')
            except:
                ws1.cell(row, 4, '测试失败').fill = red_fill
                report_result.append('fail')

            row += 1

        # 调用保存数据的方法
        wb.save('../data/mtxshop.xlsx')
        print('保存成功')
        ws2 = wb['测试用例集合']
        # 先获取case在整个case_list里面的索引值，+2就可以知道他的行数
        # 因为读取excel就是按照用例的行数一个一个去读取的
        index = case_list.index(case)
        print(f'index的值是{index}')
        if 'fail' in report_result:
            ws2.cell(index+2, 3, '测试失败').fill = red_fill
        else:
            ws2.cell(index+2, 3, '测试成功').fill = green_fill
            print('我被执行了')
        time.sleep(3)
        # 调用保存数据的方法
        wb.save('../data/mtxshop.xlsx')

        print('保存成功')
    @classmethod
    def tearDownClass(cls):
        cls.kwFunc.base_quit_driver()


if __name__ == '__main__':
    unittest.main()