# @Author: 夭夭
# @Time: 五月 28, 2020


import openpyxl
import os
import sys

o_path = os.path.abspath(os.path.join(os.getcwd(),"../"))
sys.path.append(o_path)

from util.logger import Logger

logger = Logger().getlog()


class CreateExcel:
    filename = ""

    def __init__(self, filename):
        self.filename = filename

    def get_book(self):

        base_dir = str(os.path.dirname(os.path.dirname(__file__)))
        base_dir = base_dir.replace('\\', '/')
        file_path = base_dir + '/data/' + self.filename
        wb = openpyxl.load_workbook(file_path)
        return wb

    def read_case(self):
        wb = self.get_book()
        ws = wb['测试用例集合']
        case_list = []
        for i in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            cell = [cell.value for cell in i]
            case_list.append(cell)

        return case_list


    def read2(self):
        wb = self.get_book()
        ws = wb['登录']

        case_list = []
        for i in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
            cell = [cell.value for cell in i]
            print(cell)

        return case_list


if __name__ == "__main__":
    filezpath = "mtxshop.xlsx"
    at = CreateExcel(filezpath)
    book = at.get_book()
    # case_list = at.read_case()
    case_list = at.read2()
    print(case_list)