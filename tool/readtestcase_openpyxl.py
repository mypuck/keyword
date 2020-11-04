# @Author: 夭夭
# @Time: 五月 28, 2020

import openpyxl
import os
import sys

from keywordLibrary.keywordFunction_fanshe import KeyWord
from tool.element import ELEMENT

o_path = os.path.abspath(os.path.join(os.getcwd(), "../"))
sys.path.append(o_path)


class CreateExcel:
    filename = ""

    def __init__(self, filename):
        self.filename = filename
        self.base_dir = str(os.path.dirname(os.path.dirname(__file__)))
        self.file_path = self.base_dir + '/data/' + self.filename


    def get_book(self):
        '''
        获取excel这个工作簿
        :return:
        '''

        wb = openpyxl.load_workbook(self.file_path)
        return wb

    def read_case(self):
        '''
        从测试用例集合读取出所有的case
        :return:
        '''
        wb = self.get_book()
        ws = wb['测试用例集合']
        case_list = []
        for i in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            cell = [cell.value for cell in i]
            case_list.append(cell)


        return case_list

    def read_case_y_n(self):
        '''
        如果y那么case执行
        如果n那么case不执行
        :return:
        '''
        wb = self.get_book()
        ws = wb['测试用例集合']
        case_list = []
        for i in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):
            cell = [cell.value for cell in i]
            print(cell)
            if cell[1] == 'y':
                case_list.append([cell[0]])
            else:
                cell[2]=''

        return case_list

    def read_element(self):
        '''
        获取测试元素库的数据
        :return:
        '''
        # ELEMENT ={}
        wb = self.get_book()
        ws = wb['测试元素库']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            col_value = [cell.value for cell in row]
            # 登录链接 = xpath,//div[@class='member-login']/a[text()='登录']
            ELEMENT[col_value[0]] = (col_value[1], col_value[2])
            print('ELEMENT的值是', ELEMENT)

        # return ELEMENT





if __name__ == "__main__":
    filezpath = "mtxshop.xlsx"
    at = CreateExcel(filezpath)
    case_list = at.read_case_y_n()
    print(case_list)
